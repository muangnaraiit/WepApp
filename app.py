from flask import Flask, render_template, request, redirect, jsonify, send_file
import os
import re
from io import BytesIO
from urllib.parse import unquote, urlencode
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


app = Flask(__name__)
DATABASE_URL = os.getenv("DATABASE_URL")


def get_db():
    if not DATABASE_URL:
        raise Exception("ไม่พบค่า DATABASE_URL")
    return psycopg2.connect(DATABASE_URL, sslmode="require")


def ensure_device_extra_columns():
    """
    เตรียมฐานข้อมูลให้รองรับฟิลด์ใหม่
    - เพิ่ม IP / วันประกัน / จุดประจำการ
    - ปลดล็อก MNH ให้เว้นว่างได้
    - ลบ unique constraint/index ของ MNH เพื่อให้ MNH ซ้ำได้
    """
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        ALTER TABLE devices
        ADD COLUMN IF NOT EXISTS station VARCHAR(100),
        ADD COLUMN IF NOT EXISTS ip_address VARCHAR(50),
        ADD COLUMN IF NOT EXISTS warranty_start DATE,
        ADD COLUMN IF NOT EXISTS warranty_end DATE
    """)

    # อนุญาตให้ MNH เป็นค่าว่างได้
    cur.execute("""
        ALTER TABLE devices
        ALTER COLUMN mnh DROP NOT NULL
    """)

    # ลบ UNIQUE constraint ที่ผูกกับคอลัมน์ mnh ถ้ามี
    cur.execute("""
        SELECT conname
        FROM pg_constraint
        WHERE conrelid = 'devices'::regclass
          AND contype = 'u'
          AND conkey = ARRAY[
              (
                  SELECT attnum
                  FROM pg_attribute
                  WHERE attrelid = 'devices'::regclass
                    AND attname = 'mnh'
              )
          ]
    """)
    constraints = cur.fetchall()
    for (constraint_name,) in constraints:
        cur.execute(f'ALTER TABLE devices DROP CONSTRAINT IF EXISTS "{constraint_name}"')

    # ลบ UNIQUE index ที่สร้างไว้กับ mnh โดยตรง ถ้ามี
    cur.execute("""
        SELECT indexname
        FROM pg_indexes
        WHERE tablename = 'devices'
          AND indexdef ILIKE '%UNIQUE%'
          AND indexdef ILIKE '%(mnh)%'
    """)
    indexes = cur.fetchall()
    for (index_name,) in indexes:
        cur.execute(f'DROP INDEX IF EXISTS "{index_name}"')

    conn.commit()
    cur.close()
    conn.close()


def format_device_row(row):
    """แปลงข้อมูล date ให้เป็น yyyy-mm-dd เพื่อใช้กับ input type=date และ JSON"""
    if not row:
        return None

    data = dict(row)

    for key in ("warranty_start", "warranty_end"):
        value = data.get(key)
        if hasattr(value, "isoformat"):
            data[key] = value.isoformat()
        elif value is None:
            data[key] = ""

    return data


def get_search_value():
    """รับค่าค้นหาจากทั้ง query string และ form เพื่อให้ค้างค่าหลังบันทึก/อัปเดต/ลบ"""
    return (
        request.form.get("search", "")
        or request.args.get("search", "")
        or ""
    ).strip()


def redirect_home_with_search(search_value):
    """กลับหน้าแรกพร้อมคงค่าค้นหาเดิมไว้ ถ้ามีค่า"""
    search_value = (search_value or "").strip()

    if search_value:
        return redirect("/?" + urlencode({"search": search_value}))

    return redirect("/")


def clean_mnh(value):
    """
    ทำความสะอาด MNH
    - เว้นว่างได้
    - ถ้าวาง URL barcode จะตัด prefix ออก
    - ถ้ามีอักขระแปลก ๆ ให้คืนค่าว่างเพื่อไม่ให้ข้อมูลขยะเข้า DB
    """
    if not value:
        return ""

    value = unquote(str(value).strip())

    prefix = "https://www.dsmetsmart.com/dsmet_hos/test_barcode.php?id="
    if value.lower().startswith(prefix.lower()):
        value = value[len(prefix):]

    value = value.upper().strip()

    if value == "":
        return ""

    if not re.match(r"^[A-Z0-9\-]+$", value):
        return ""

    return value


def excel_text(value):
    """
    แปลงค่าทุกช่องให้เป็น Text สำหรับ Excel
    - None เป็นค่าว่าง
    - ค่าอื่นแปลงเป็น string
    - กัน Excel ตีความเป็นสูตร ถ้าขึ้นต้นด้วย = + - @
    """
    if value is None:
        return ""

    value = str(value)

    if value.startswith(("=", "+", "-", "@")):
        return "'" + value

    return value


ORDER_SQL = """
    ORDER BY
        station ASC NULLS LAST,
        computer_name ASC NULLS LAST,

        CASE device_type
            WHEN 'PC' THEN 1
            WHEN 'AIO' THEN 2
            WHEN 'Notebook' THEN 3
            WHEN 'Monitor' THEN 4
            WHEN 'Printer' THEN 5
            WHEN 'Scanner' THEN 6
            WHEN 'UPS' THEN 7
            WHEN 'Phone' THEN 8
            WHEN 'Switch Hub' THEN 9
            WHEN 'Projector' THEN 10
            ELSE 99
        END ASC,

        model ASC NULLS LAST,
        serial_number ASC NULLS LAST,
        mnh ASC NULLS LAST
"""


@app.route("/")
def home():
    ensure_device_extra_columns()

    search_raw = request.args.get("search", "").strip()
    search = clean_mnh(search_raw) or search_raw

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    if search:
        keyword = f"%{search}%"
        cur.execute(f"""
            SELECT *
            FROM devices
            WHERE COALESCE(mnh, '') ILIKE %s
               OR COALESCE(station, '') ILIKE %s
               OR COALESCE(device_type, '') ILIKE %s
               OR COALESCE(model, '') ILIKE %s
               OR COALESCE(serial_number, '') ILIKE %s
               OR COALESCE(computer_name, '') ILIKE %s
               OR COALESCE(ip_address, '') ILIKE %s
               OR COALESCE(TO_CHAR(warranty_start, 'YYYY-MM-DD'), '') ILIKE %s
               OR COALESCE(TO_CHAR(warranty_end, 'YYYY-MM-DD'), '') ILIKE %s
            {ORDER_SQL}
        """, (keyword, keyword, keyword, keyword, keyword, keyword, keyword, keyword, keyword))
    else:
        cur.execute(f"""
            SELECT *
            FROM devices
            {ORDER_SQL}
        """)

    rows = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT model
        FROM devices
        WHERE model IS NOT NULL
          AND model <> ''
        ORDER BY model
    """)
    models = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT computer_name
        FROM devices
        WHERE computer_name IS NOT NULL
          AND computer_name <> ''
        ORDER BY computer_name
    """)
    computer_names = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "index.html",
        rows=rows,
        models=models,
        computer_names=computer_names,
        search=search
    )


@app.route("/api/device/<path:mnh>")
def api_device(mnh):
    ensure_device_extra_columns()

    mnh = clean_mnh(mnh)

    if not mnh:
        return jsonify({"found": False, "items": []})

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # MNH ซ้ำได้ จึงต้องคืนทุกรายการให้หน้าเว็บเลือก
    cur.execute(f"""
        SELECT *
        FROM devices
        WHERE mnh = %s
        {ORDER_SQL}
    """, (mnh,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    items = [format_device_row(row) for row in rows]

    if items:
        return jsonify({
            "found": True,
            "data": items[0],
            "items": items
        })

    return jsonify({"found": False, "items": []})


@app.route("/save", methods=["POST"])
def save():
    ensure_device_extra_columns()

    current_search = get_search_value()

    edit_id_raw = request.form.get("edit_id", "").strip()

    # MNH ไม่บังคับกรอก และซ้ำได้
    mnh = clean_mnh(request.form.get("mnh", ""))
    station = request.form.get("station", "").strip()
    device_type = request.form.get("device_type", "").strip()
    model = request.form.get("model", "").strip()
    serial_number = request.form.get("serial_number", "").strip()
    computer_name = request.form.get("computer_name", "").strip()
    ip_address = request.form.get("ip_address", "").strip()
    warranty_start = request.form.get("warranty_start") or None
    warranty_end = request.form.get("warranty_end") or None

    if device_type == "PC":
        model = "PC"
        serial_number = "PC"

    edit_id = None
    if edit_id_raw.isdigit():
        edit_id = int(edit_id_raw)

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        if edit_id:
            # โหมดแก้ไข: แก้ตาม id เท่านั้น เพื่อให้ MNH ซ้ำได้และไม่ไปทับรายการอื่น
            cur.execute("""
                UPDATE devices
                SET
                    mnh = %s,
                    station = %s,
                    device_type = %s,
                    model = %s,
                    serial_number = %s,
                    computer_name = %s,
                    ip_address = %s,
                    warranty_start = %s,
                    warranty_end = %s
                WHERE id = %s
            """, (
                mnh,
                station,
                device_type,
                model,
                serial_number,
                computer_name,
                ip_address,
                warranty_start,
                warranty_end,
                edit_id
            ))
        else:
            # โหมดเพิ่มข้อมูล: INSERT ใหม่เสมอ เพื่อรองรับ MNH ซ้ำ/ว่าง
            cur.execute("""
                INSERT INTO devices (
                    mnh,
                    station,
                    device_type,
                    model,
                    serial_number,
                    computer_name,
                    ip_address,
                    warranty_start,
                    warranty_end
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                mnh,
                station,
                device_type,
                model,
                serial_number,
                computer_name,
                ip_address,
                warranty_start,
                warranty_end
            ))

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()

    return redirect_home_with_search(current_search)


@app.route("/delete/<int:id>", methods=["POST"])
def delete_device(id):
    current_search = get_search_value()

    conn = get_db()
    cur = conn.cursor()

    cur.execute("DELETE FROM devices WHERE id = %s", (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect_home_with_search(current_search)


@app.route("/export_excel")
def export_excel():
    ensure_device_extra_columns()

    search_raw = request.args.get("search", "").strip()
    search = clean_mnh(search_raw) or search_raw

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    if search:
        keyword = f"%{search}%"
        cur.execute(f"""
            SELECT
                station,
                computer_name,
                device_type,
                model,
                serial_number,
                mnh,
                ip_address,
                warranty_start,
                warranty_end
            FROM devices
            WHERE COALESCE(mnh, '') ILIKE %s
               OR COALESCE(station, '') ILIKE %s
               OR COALESCE(device_type, '') ILIKE %s
               OR COALESCE(model, '') ILIKE %s
               OR COALESCE(serial_number, '') ILIKE %s
               OR COALESCE(computer_name, '') ILIKE %s
               OR COALESCE(ip_address, '') ILIKE %s
               OR COALESCE(TO_CHAR(warranty_start, 'YYYY-MM-DD'), '') ILIKE %s
               OR COALESCE(TO_CHAR(warranty_end, 'YYYY-MM-DD'), '') ILIKE %s
            {ORDER_SQL}
        """, (keyword, keyword, keyword, keyword, keyword, keyword, keyword, keyword, keyword))
    else:
        cur.execute(f"""
            SELECT
                station,
                computer_name,
                device_type,
                model,
                serial_number,
                mnh,
                ip_address,
                warranty_start,
                warranty_end
            FROM devices
            {ORDER_SQL}
        """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "รายการอุปกรณ์"

    headers = [
        "จุดประจำการ",
        "ชื่อเครื่อง",
        "ประเภท",
        "รุ่น",
        "Serial Number",
        "MNH",
        "IP Address",
        "วันเริ่มต้นประกัน",
        "วันหมดประกัน"
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.number_format = "@"

    for row in rows:
        ws.append([
            excel_text(row.get("station")),
            excel_text(row.get("computer_name")),
            excel_text(row.get("device_type")),
            excel_text(row.get("model")),
            excel_text(row.get("serial_number")),
            excel_text(row.get("mnh")),
            excel_text(row.get("ip_address")),
            excel_text(row.get("warranty_start")),
            excel_text(row.get("warranty_end")),
        ])

    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = excel_text(cell.value)

    column_widths = {
        "A": 22,
        "B": 24,
        "C": 18,
        "D": 30,
        "E": 24,
        "F": 20,
        "G": 18,
        "H": 20,
        "I": 20,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "devices.xlsx"
    if search:
        safe_search = re.sub(r"[^A-Za-z0-9ก-๙_-]+", "_", search)
        filename = f"devices_{safe_search}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
